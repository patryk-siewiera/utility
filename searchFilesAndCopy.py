import shutil
from openpyxl import load_workbook
import os
import glob
from datetime import datetime


# if app don't work
# pip install openpyxl


def app():
    # *********** EDIT ME
    searchFolder = r"C:\Users\sievr\Downloads\KKCE\wnioski materiałowe"
    destinationPath = r"C:\Users\sievr\Downloads\KKCE\solution\here_paste_solutions"
    xlsName = r"C:\Users\sievr\Downloads\KKCE\solution\excelData.xlsx"
    excelMaxColumn__Y = 4
    excelMaxRow__X = 8
    preserveOriginalFilename = False

    # --------------------------------
    searchSubfolders = "\**"
    origin = searchFolder + searchSubfolders
    destination = os.path.join(destinationPath, nowCurrentTime())

    readAndPrintInitValues(origin, destination, xlsName, excelMaxColumn__Y, excelMaxRow__X)
    xls = readXlsAndReturnValues(xlsName, excelMaxColumn__Y, excelMaxRow__X)
    manipulateXls(xls, destination, origin, preserveOriginalFilename)


def manipulateXls(xls, destination, origin, preserveOriginalFilename):
    for item in xls:
        folderName = item[0]
        keywordsTemp = item[1:]
        destinationTemp = (os.path.join(destination, folderName))
        copyAllFiles(origin, destinationTemp, keywordsTemp, folderName, preserveOriginalFilename)


def nowCurrentTime():
    nowTime = datetime.now()
    nowTime = nowTime.strftime("%Y-%m-%d__%H-%M-%S")
    return str(nowTime)


def readAndPrintInitValues(origin, destination, xlsName, excelMaxColumn, excelMaxRow):
    print("\n*** YOUR VALUES ***")
    print(
        "*** Excel filename\n\t" + xlsName + " \t\t[MAX_columns: " + str(excelMaxColumn) + ", MAX_rows: " + str(
            excelMaxRow) + "]")
    print("*** source folder\n\t" + origin)
    print("*** destination\n\t" + destination)


def readXlsAndReturnValues(xlsName, excelMaxColumn, excelMaxRow):
    wb = load_workbook(xlsName)
    sheet = wb.active
    rows_iter = sheet.iter_rows(max_col=excelMaxColumn, max_row=excelMaxRow)
    allValuesFromXLS = [[cell.value for cell in list(row)] for row in rows_iter]
    return allValuesFromXLS


def createFolderIfNotExist(pathToNewFolder):
    if not os.path.exists(pathToNewFolder):
        os.makedirs(pathToNewFolder)
        print("\n++ Folders created! \n\t", pathToNewFolder, "\n")
    else:
        print("\n-- Folder already exist,  \n\t", pathToNewFolder, "\n")


def filterArray(array, keyWords, path):
    # array and keyWords will be reduced to lowercase
    array = [each_string.lower() for each_string in array]
    keyWords = [each_keyWords.lower() for each_keyWords in keyWords]
    for key in keyWords:
        array = [item for item in array if key in item]
    if len(array) == 0 or not os.path.isfile(path):
        return False
    else:
        return True


def copyAllFiles(origin, destination, keyWords, folderName, preserveOriginalFilename):
    listOfAllFilesFullPath = []
    keyWords = [x for x in keyWords if x is not None]
    print("\n------ FOLDER NAME:\t", folderName)
    print("------ KEYWORDS:\t", keyWords)
    print("\n++ Files found: \t")
    for f in glob.glob(origin, recursive=True):
        if (filterArray([os.path.basename(f)], keyWords, f)):
            listOfAllFilesFullPath.append(f)
            print("\t" + os.path.basename(f))
    if (len(listOfAllFilesFullPath) == 0):
        print("\t-- !! -- \t List is empty\n\n\n")
        return False

    try:
        createFolderIfNotExist(destination)
        id = 0
        for fileName in listOfAllFilesFullPath:
            if os.walk(fileName):
                print(os.path.basename(fileName))
                shutil.copy(fileName, destination)
                fileTempPath = (os.path.join(destination, os.path.basename(fileName)))
                keyWordBuilder = "_".join(keyWords)
                extension = (os.path.splitext(fileName)[1])
                if not (preserveOriginalFilename):
                    if (id == 0):
                        os.rename(fileTempPath, destination + "/" + keyWordBuilder + extension)
                    else:
                        os.rename(fileTempPath, destination + "/" + keyWordBuilder + "__(" + str(id) + ")" + extension)
                else:
                    if (id == 0):
                        os.rename(fileTempPath,
                                  destination + "/" + str(os.path.basename(
                                      fileName))
                                  + "___" + keyWordBuilder + "_____" + extension)
                    else:
                        os.rename(fileTempPath,
                                  destination + "/" + str(os.path.basename(
                                      fileName))
                                  + "_____" + keyWordBuilder + "__(" + str(id) + ")" + extension)

                id = id + 1
        print("++ Files copied successfully \n\n\n")
        id = 0
    except:
        print("\n--!!!--\t Fail during copying files")


app()
